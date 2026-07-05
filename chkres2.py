import openpyxl, os
p=os.path.join('C:'+os.sep,'Users','Administrator','Desktop')
wb=openpyxl.load_workbook(os.path.join(p,'水池_结果.xlsx'))
ws=wb.active
print('rows:',ws.max_row)
for r in range(1, ws.max_row+1):
    vals=[str(ws.cell(r,c).value or'') for c in range(1,8)]
    print('|'.join(vals))
