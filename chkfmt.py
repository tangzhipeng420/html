import openpyxl, os
p=os.path.join('C:'+os.sep,'Users','Administrator','Desktop')
wb=openpyxl.load_workbook(os.path.join(p,'龙虾查的.xlsx'))
ws=wb.active
print('rows:',ws.max_row,'cols:',ws.max_column)
for r in range(1, min(6, ws.max_row+1)):
    row=[str(ws.cell(r,c).value or '') for c in range(1, ws.max_column+1)]
    print('|'.join(row))
