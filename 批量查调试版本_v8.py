#!/usr/bin/env python3
"""
BOSS水池批量查询 v8
- 每个号码独立查询，查完关闭Tab
- 无效号码等5秒，有效号码等60-240秒随机
- 营销活动取"营销包名称"列（tds[3]），不重复写入
- Tab name加行号后缀避免复用

用法：python3 批量查调试版本_v8.py [起始行] [结束行] [每日上限]
默认起始行: 126, 每日上限: 130
"""
import asyncio, json, urllib.request, websockets, time, re, sys, random
from datetime import datetime

HOST = "172.27.144.1"
PORT = 19228
POOL_XLSX = "/mnt/c/Users/Administrator/Desktop/水池.xlsx"
DAILY_LIMIT = 130

EQ_URL = "/ordercentre/ordercentre?service=page/oc.person.cs.fusion.Equitypoolquery&listener=onInitBusi&MENU_ID=FUSE20210917"
MB_URL = "/agentcentre/agentcentre?service=page/person.other.vc.MultiShareMangement&MENU_ID=FUSEQ20180427"
MK_URL = "/agentcentre/agentcentre?service=page/person.saleActive.vc.QuerySaleActive&listener=init&MENU_ID=FUSE5004"

COL_PHONE=1; COL_CUST=2; COL_PLAN=3; COL_GOTONE=4
COL_ARPU_M=5; COL_ARPU_Y=6; COL_EQUITY=7; COL_MULTI=8

import openpyxl

def save_xlsx(wb):
    for _ in range(5):
        try:
            wb.save(POOL_XLSX)
            return True
        except:
            time.sleep(1)
    return False

def get_val(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None: return ''
    s = str(v).strip()
    return '' if s in ('None', '') else s

async def main():
    start_line = int(sys.argv[1]) if len(sys.argv) > 1 else 126
    end_line = int(sys.argv[2]) if len(sys.argv) > 2 else None
    limit = int(sys.argv[3]) if len(sys.argv) > 3 else DAILY_LIMIT

    print(f"{'='*60}")
    print(f"  BOSS水池批量查询 v8")
    print(f"  起始行: {start_line}" + (f"  结束行: {end_line}" if end_line else ""))
    print(f"  每日上限: {limit}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(POOL_XLSX)
    ws = wb.active
    print(f"  Excel总行数: {ws.max_row}")

    if end_line:
        todo = [(r, get_val(ws, r, COL_PHONE)) for r in range(start_line, end_line+1) if get_val(ws, r, COL_PHONE)]
    else:
        todo = [(r, get_val(ws, r, COL_PHONE)) for r in range(1, ws.max_row+1)
                if get_val(ws, r, COL_PHONE) and not get_val(ws, r, COL_CUST) and r >= start_line]

    if not todo:
        print("  没有需要查询的行")
        return

    print(f"  待查询: {len(todo)} 个")
    for r,p in todo[:5]:
        print(f"    行{r}: {p}")
    if len(todo) > 5: print(f"    ... 共{len(todo)}个")

    r = json.loads(urllib.request.urlopen(f"http://{HOST}:{PORT}/json", timeout=10).read())
    ws_url = r[0]['webSocketDebuggerUrl']

    def mid(): return int(time.time_ns() % 100000)
    valid = 0
    invalid = 0

    async with websockets.connect(ws_url, max_size=2**24, ping_interval=30) as ws_:

        async def js(code):
            try:
                await ws_.send(json.dumps({'id':mid(),'method':'Runtime.evaluate','params':{'expression':code,'returnByValue':True}}))
                d = json.loads(await asyncio.wait_for(ws_.recv(), timeout=30))
                return d.get('result',{}).get('result',{}).get('value')
            except:
                return None

        async def home():
            await js("""(function(){
                var pops=document.querySelectorAll('.c_dialog,.mask,.overlay,.c_guide,.c_popupBg,.dialog');
                for(var i=0;i<pops.length;i++)pops[i].style.display='none';
                var btns=document.querySelectorAll('input[value="取消"],button');
                for(var i=0;i<btns.length;i++){
                    if(btns[i].textContent.trim()==='取消'||btns[i].value==='取消') btns[i].click();
                }
                var fs=document.querySelectorAll('iframe');
                for(var i=0;i<fs.length;i++){
                    if(fs[i].id==='navframe_def') fs[i].style.display='';
                    else if(fs[i].id.indexOf('navframe_')===0) fs[i].style.display='none';
                }
            })()""")
            await asyncio.sleep(1)

        async def close_tab(fid):
            if not fid:
                return
            num = fid.replace('navframe_', '')
            await js(f"(function(){{var btn=document.getElementById('navtab_close_'+'{num}');if(btn)btn.click();}})()")
            await asyncio.sleep(0.5)

        async def open_page(name, url):
            await js(f"(function(){{openNavByUrl('{name}','{url}','',{{}});}})()")
            for _ in range(30):
                await asyncio.sleep(0.5)
                fid = await js("""(function(){
                    var fs=document.querySelectorAll('iframe');
                    for(var i=fs.length-1;i>=0;i--){
                        if(fs[i].style.display!=='none' && fs[i].id.indexOf('navframe_')===0 && fs[i].id!=='navframe_def'){
                            return fs[i].id;
                        }
                    }
                    return '';
                })()""")
                if fid:
                    return fid
            return None

        for idx, (excel_row, phone) in enumerate(todo):
            if valid >= limit:
                print(f"\n🛑 达到上限 {limit}")
                break

            now = datetime.now().strftime('%H:%M:%S')
            print(f"\n[{now}] [{idx+1}/{len(todo)}] 📱 {phone} (行{excel_row})")

            wb = openpyxl.load_workbook(POOL_XLSX)
            ws = wb.active

            try:
                # ═══ 1. 基础信息 ═══
                await home()
                print(f"  [1] 基础信息...")
                await js(f"(function(){{var inp=document.querySelector('#verifyInput');if(inp){{inp.value='{phone}';inp.focus();}}var ev=new KeyboardEvent('keydown',{{key:'Enter',code:'Enter',keyCode:13,which:13}});window.event=ev;if(typeof enterAction==='function')enterAction(qryAccNumber);}})()")

                name = acc_phone = None
                invalid_flag = False
                for s in range(30):
                    await asyncio.sleep(1)
                    cust_el = await js("(function(){var el=document.querySelector('#cust_info');return el?el.innerText||'found':'';})()")
                    if cust_el and '无此号码' in cust_el:
                        invalid_flag = True
                        break
                    name = await js("(function(){var el=document.querySelector('#cust_info .CUST_NAME2 .value');return el?el.textContent.trim():'';})()")
                    acc_phone = await js("(function(){var el=document.querySelector('#cust_info .ACCESS_NUMBER .value');return el?el.textContent.trim():'';})()")
                    if name and acc_phone == phone:
                        break

                if invalid_flag or not name or acc_phone != phone:
                    print(f"  ❌ 无效号码")
                    ws.cell(row=excel_row, column=COL_CUST).value = '无效号码'
                    save_xlsx(wb)
                    invalid += 1
                    print(f"  📊 ✅{valid} ❌{invalid} | 额度:{limit-valid}")
                    print(f"  ⏳ 等待5s（无效号码）...")
                    await asyncio.sleep(5)
                    continue

                info_raw = await js("""(function(){var r={};var lis=document.querySelectorAll('#cust_info li');for(var i=0;i<lis.length;i++){var li=lis[i];var cls=li.className;var v=li.querySelector('.value');var t=v?v.textContent.trim():'';if(cls&&t&&cls!=='MSISDN'){r[cls]=t;}}return JSON.stringify(r);})()""")
                plan = gotone = arpu_m = arpu_y = ''
                if info_raw:
                    raw = json.loads(info_raw)
                    plan = raw.get('OFFER_NAME','')
                    gotone = raw.get('GOTONE_MONTH_LEVEL','')
                    arpu_m = raw.get('GOTONE_MONTH_ARPU','')
                    arpu_y = raw.get('GOTONE_YEAR_ARPU','')
                print(f"  ✅ {name} | {plan} | {gotone} | ARPU月:{arpu_m} 年:{arpu_y}")

                # ═══ 2. 权益金 ═══
                equity = ''
                print(f"  [2] 权益金...")
                eq_fid = await open_page('eq'+str(excel_row), EQ_URL)
                if eq_fid:
                    await js(f"(function(){{var f=document.querySelector('#{eq_fid}');if(!f||!f.contentDocument)return;var doc=f.contentDocument;var inp=doc.querySelector('#ACCESS_NUMBER');if(inp)inp.value='{phone}';var w=f.contentWindow;if(w.checkBaseQuery){{w.checkBaseQuery();}}else{{var btns=doc.querySelectorAll('button');for(var i=0;i<btns.length;i++){{if(btns[i].textContent.indexOf('执行')>=0){{btns[i].click();break;}}}}}}}})()")
                    await asyncio.sleep(5)
                    text = await js(f"(function(){{var f=document.querySelector('#{eq_fid}');if(!f||!f.contentDocument)return'';return f.contentDocument.body.innerText||'';}})()")
                    m = re.search(r'权益金余额[：:]\s*([\d,]+\.?\d*)', text) if text else None
                    equity = m.group(1) if m else '0'
                print(f"  权益金: {equity if equity else '-'}")
                await close_tab(eq_fid)

                # ═══ 3. 多人保底 ═══
                multi = ''
                print(f"  [3] 多人保底...")
                mb_fid = await open_page('mb'+str(excel_row), MB_URL)
                if mb_fid:
                    await js(f"(function(){{var f=document.querySelector('#{mb_fid}');if(!f||!f.contentDocument)return;var doc=f.contentDocument;var pops=doc.querySelectorAll('.c_dialog,.mask,.overlay,.c_guide,.c_popupBg');for(var i=0;i<pops.length;i++)pops[i].remove();var inp=doc.querySelector('#ACCESS_NUMBER');if(inp){{inp.disabled=false;inp.classList.remove('e_dis');inp.value='{phone}';}}var w=f.contentWindow;if(w.checkBaseQueryWithOutVerify){{w.checkBaseQueryWithOutVerify();}}else{{var btn=doc.querySelector('#qryWithOutVerifybtn');if(btn)btn.click();}}}})()")
                    await asyncio.sleep(5)
                    await js(f"(function(){{var f=document.querySelector('#{mb_fid}');if(!f||!f.contentDocument)return;var doc=f.contentDocument;var els=doc.querySelectorAll('*');for(var i=0;i<els.length;i++){{var el=els[i];var txt=el.textContent||'';if(el.tagName==='DIV'&&txt.trim()==='✕'&&el.parentElement&&el.parentElement.innerText.indexOf('选出图中')>=0){{el.click();break;}}}}}})()")
                    await asyncio.sleep(2)
                    nt = await js(f"(function(){{var f=document.querySelector('#{mb_fid}');if(!f||!f.contentDocument)return'';var doc=f.contentDocument;var el=doc.querySelector('#ACCESS_NUM_TYPE');return el?el.textContent.trim():'';}})()")
                    multi = nt if nt else '无'
                print(f"  多人保底: {multi if multi else '-'}")
                await close_tab(mb_fid)

                # ═══ 4. 营销活动 ═══
                print(f"  [4] 营销活动...")
                mk_list = []
                mk_name = f'mk{excel_row}'
                mk_fid = await open_page(mk_name, MK_URL)
                if mk_fid:
                    await js(f"(function(){{var f=document.querySelector('#{mk_fid}');if(!f||!f.contentDocument)return;var doc=f.contentDocument;var inp=doc.querySelector('#ACCESS_NUMBER');if(inp)inp.value='{phone}';var btns=doc.querySelectorAll('button');for(var i=0;i<btns.length;i++){{if(btns[i].textContent.indexOf('执行查询')>=0){{btns[i].click();break;}}}}}})()")
                    await asyncio.sleep(8)
                    for pg in range(30):
                        raw = await js(f"(function(){{var f=document.querySelector('#{mk_fid}');if(!f||!f.contentDocument)return'__dead__';var doc=f.contentDocument;var tbl=doc.querySelector('#mytable_body_table');if(!tbl)return'__dead__';var n=[];var rs=tbl.querySelectorAll('tr');for(var j=1;j<rs.length;j++){{var tds=rs[j].querySelectorAll('td');var mk=tds[3]?tds[3].textContent.trim():'';if(mk)n.push(mk);}}return JSON.stringify(n);}})()")
                        print(f"    p{pg+1}: {len(json.loads(raw)) if raw not in ('__dead__','[]') else raw} 条")
                        if raw == '__dead__':
                            break
                        try:
                            names = json.loads(raw)
                            mk_list.extend(names)
                        except:
                            pass
                        pi = await js(f"(function(){{var f=document.querySelector('#{mk_fid}');if(!f||!f.contentDocument)return'';var t=f.contentDocument.body.innerText;var m=t.match(/(\\d+)\\/(\\d+)/);return m?m[1]+'/'+m[2]:'no_match';}})()")
                        if not pi or pi == 'no_match':
                            break
                        m = re.match(r'(\d+)/(\d+)', pi)
                        if m and int(m.group(1)) >= int(m.group(2)):
                            break
                        await js(f"(function(){{var f=document.querySelector('#{mk_fid}');if(!f||!f.contentDocument)return;var n=f.contentDocument.querySelector('#mytablenavbar_next');if(n)n.click();}})()")
                        await asyncio.sleep(3)
                print(f"  营销包: {len(mk_list)}条")
                await close_tab(mk_fid)

                # ═══ 写回固定列 ═══
                ws.cell(row=excel_row, column=COL_CUST).value = name
                ws.cell(row=excel_row, column=COL_PLAN).value = plan
                ws.cell(row=excel_row, column=COL_GOTONE).value = gotone
                ws.cell(row=excel_row, column=COL_ARPU_M).value = arpu_m
                ws.cell(row=excel_row, column=COL_ARPU_Y).value = arpu_y
                ws.cell(row=excel_row, column=COL_EQUITY).value = equity if equity else '0'
                ws.cell(row=excel_row, column=COL_MULTI).value = multi if multi else '无'

                # ═══ 营销活动展开 ═══
                if mk_list:
                    header_map = {}
                    max_col = ws.max_column
                    for c in range(9, max_col + 1):
                        hv = ws.cell(row=1, column=c).value
                        if hv:
                            header_map[str(hv).strip()] = c
                    next_col = max_col + 1
                    for mk_name in mk_list:
                        mk_name = mk_name.strip()
                        if not mk_name:
                            continue
                        if mk_name in header_map:
                            ws.cell(row=excel_row, column=header_map[mk_name]).value = '是'
                        else:
                            ws.cell(row=1, column=next_col).value = mk_name
                            ws.cell(row=excel_row, column=next_col).value = '是'
                            header_map[mk_name] = next_col
                            next_col += 1
                            print(f"    ➕ 新增列: {mk_name}")
                save_xlsx(wb)
                print(f"  💾 已保存")
                valid += 1

            except Exception as e:
                print(f"  ⚠️ 异常: {e}")
                import traceback
                traceback.print_exc()
                try:
                    wb = openpyxl.load_workbook(POOL_XLSX)
                    wb.active.cell(row=excel_row, column=COL_CUST).value = '无效号码'
                    save_xlsx(wb)
                except:
                    pass
                invalid += 1

            print(f"  📊 ✅有效{valid} ❌无效{invalid} | 额度:{limit-valid}")
            delay = random.uniform(60, 240)
            print(f"  ⏳ 等待{delay:.0f}s（有效号码）...")
            await asyncio.sleep(delay)

    print(f"\n{'='*60}")
    print(f"  🏁 完成 ✅{valid} ❌{invalid}")
    print(f"{'='*60}")

asyncio.run(main())
