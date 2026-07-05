import openpyxl, os
p=os.path.join('C:'+os.sep,'Users','Administrator','Desktop')
wb=openpyxl.load_workbook(os.path.join(p,'水池_结果.xlsx'))
ws=wb.active
print('rows:',ws.max_row)
for r in range(1, ws.max_row+1):
    ph=str(ws.cell(r,1).value or '')
    res=str(ws.cell(r,2).value or '')[:200]
    st=str(ws.cell(r,3).value or '')
    print(f'#{r}: {ph} | {res} | {st}')
